"""Vérification de bout en bout de la page Données via AppTest.

Isolation : `lib.project.PREFS_PATH` monkeypatché (comme
`tests/test_app_projet.py`) pour ne jamais toucher `Path.home()` réel ni
`DEFAULT_PROJECTS_ROOT` (`D:\\EthoFlow\\projects`, un nom de dossier
littéral et relatif sur ce runner macOS — piège déjà tombé deux fois).

Le projet courant est injecté directement dans `at.session_state`
(`current_project_path`), comme le fait `lib.config.set_current_project`,
plutôt qu'en pilotant le sélecteur de la page Projet : ces tests portent
sur la page Données, pas sur l'ouverture de projet.
"""
from __future__ import annotations

import sys
from pathlib import Path

import yaml
from streamlit.testing.v1 import AppTest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
from excel_templates import write_starter_excel  # noqa: E402

from lib import pipeline as PL
from lib import project as P
from lib import runner as R

REPO = Path(__file__).resolve().parent.parent
APP_PY = str(REPO / "streamlit_app" / "app.py")


def _argv_sans_conda(cmd: PL.Command) -> list[str]:
    """`conda run -n <env> python <script> <args>` -> `[sys.executable, script, *args]`."""
    argv = PL.to_argv(cmd)
    return [sys.executable] + argv[5:]


def _attendre_job_termine(project: Path, timeout: float = 20.0) -> None:
    import time
    fin = time.time() + timeout
    while time.time() < fin:
        job = R.current(project)
        if job and job.state != "running":
            return
        time.sleep(0.05)
    raise AssertionError("job toujours 'running' après timeout")


def _projet_vide(tmp_path: Path) -> Path:
    p = tmp_path / "projects" / "test-plan"
    for sub in ("raw", "cropped", "dlc-output", "vame", "results"):
        (p / "data" / sub).mkdir(parents=True)
    (p / "configs").mkdir(parents=True)
    (p / "configs" / "pipeline_config.yaml").write_text(
        yaml.safe_dump({"kind": "single"}, sort_keys=False), encoding="utf-8",
    )
    return p


def _lancer_sur_projet(tmp_path: Path, monkeypatch, projet: Path) -> AppTest:
    monkeypatch.setattr(P, "PREFS_PATH", tmp_path / "app_prefs.yaml")
    P.save_prefs({"projects_root": str(projet.parent), "models_root": str(tmp_path / "models")})
    (tmp_path / "models").mkdir(exist_ok=True)
    at = AppTest.from_file(APP_PY)
    at.session_state["current_project_path"] = str(projet)
    at.run()
    assert not at.exception, at.exception
    boutons = {b.key: b for b in at.button}
    assert "nav_donnees" in boutons, list(boutons)
    boutons["nav_donnees"].click().run()
    assert not at.exception, at.exception
    return at


def test_avec_excel_le_telechargement_est_offert(tmp_path, monkeypatch):
    projet = _projet_vide(tmp_path)
    write_starter_excel(projet / f"{projet.name}_sessions.xlsx", "single", projet.name)

    at = _lancer_sur_projet(tmp_path, monkeypatch, projet)

    # `streamlit.testing.v1` (1.57) n'a pas d'accesseur `download_button` :
    # on vérifie l'absence d'exception (le `data=excel_path.read_bytes()`
    # planterait si le fichier était introuvable) et le message qui
    # accompagne le bouton.
    assert not at.exception, at.exception
    assert any("Excel trouvé" in s.value for s in at.success), [s.value for s in at.success]


def test_sans_excel_la_page_le_dit_sans_planter(tmp_path, monkeypatch):
    projet = _projet_vide(tmp_path)
    # Pas d'Excel écrit du tout.

    at = _lancer_sur_projet(tmp_path, monkeypatch, projet)

    assert not at.exception, at.exception
    assert any("Aucun Excel trouvé" in w.value for w in at.warning), [w.value for w in at.warning]


def test_boutons_sync_desactives_sans_dossier_videos(tmp_path, monkeypatch):
    projet = _projet_vide(tmp_path)
    write_starter_excel(projet / f"{projet.name}_sessions.xlsx", "single", projet.name)

    at = _lancer_sur_projet(tmp_path, monkeypatch, projet)

    boutons = {b.key: b for b in at.button}
    assert boutons["btn_sync_apercu"].disabled
    assert boutons["btn_sync_reel"].disabled

    videos_dir = tmp_path / "videos"
    videos_dir.mkdir()
    champ = [t for t in at.text_input if t.label == "Dossier des vidéos"][0]
    champ.set_value(str(videos_dir)).run()
    assert not at.exception, at.exception

    boutons = {b.key: b for b in at.button}
    assert not boutons["btn_sync_apercu"].disabled
    assert not boutons["btn_sync_reel"].disabled


def test_upload_demande_confirmation_avant_ecrasement(tmp_path, monkeypatch):
    projet = _projet_vide(tmp_path)
    excel_path = projet / f"{projet.name}_sessions.xlsx"
    write_starter_excel(excel_path, "single", projet.name)
    contenu_original = excel_path.read_bytes()

    at = _lancer_sur_projet(tmp_path, monkeypatch, projet)

    uploader = at.file_uploader[0]
    uploader.upload("rempli.xlsx", b"pas un vrai xlsx, juste un contenu de test").run()
    assert not at.exception, at.exception

    # La confirmation est demandée : le fichier original n'a pas bougé.
    assert excel_path.read_bytes() == contenu_original
    assert any("Écraser" in w.value for w in at.warning), [w.value for w in at.warning]
    boutons = {b.key: b for b in at.button}
    assert "btn_confirmer_upload_excel" in boutons
    assert "btn_annuler_upload_excel" in boutons

    boutons["btn_confirmer_upload_excel"].click().run()
    assert not at.exception, at.exception
    assert excel_path.read_bytes() == b"pas un vrai xlsx, juste un contenu de test"


def test_sync_dry_run_reel_bout_en_bout(tmp_path, monkeypatch):
    """Aperçu (dry-run) réel : construit les commandes via lib.pipeline,
    les exécute via lib.runner (sans conda, sys.executable direct), et
    vérifie que rien n'est écrit sous data/raw/."""
    monkeypatch.setattr(R, "_argv", _argv_sans_conda)

    projet = _projet_vide(tmp_path)
    excel_path = projet / f"{projet.name}_sessions.xlsx"
    write_starter_excel(excel_path, "single", projet.name)

    # Remplir la feuille Sessions avec deux lignes, dont un id sans vidéo.
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Sessions"]
    headers = [c.value for c in ws[1]]
    id_col = headers.index("id") + 1
    ws.cell(row=2, column=id_col, value="vid-un")
    ws.cell(row=3, column=id_col, value="vid-manquante")
    wb.save(excel_path)

    videos_dir = tmp_path / "videos"
    videos_dir.mkdir()
    (videos_dir / "vid-un.mp4").write_bytes(b"\x00")
    # vid-manquante.mp4 n'existe pas exprès.

    at = _lancer_sur_projet(tmp_path, monkeypatch, projet)

    champ = [t for t in at.text_input if t.label == "Dossier des vidéos"][0]
    champ.set_value(str(videos_dir)).run()
    assert not at.exception, at.exception

    boutons = {b.key: b for b in at.button}
    boutons["btn_sync_apercu"].click().run()
    assert not at.exception, at.exception

    _attendre_job_termine(projet)
    at.run()
    assert not at.exception, at.exception

    job = R.current(projet)
    assert job.state == "succeeded", R.read_log(projet, job.job_id)
    log = R.read_log(projet, job.job_id)
    print("\n----- LOG APERÇU DRY-RUN -----\n" + log)
    assert "vid-un" in log
    assert "vid-manquante" in log
    assert "à écrire" in log

    # Dry-run : rien n'a dû être écrit sous data/raw/.
    raw = projet / "data" / "raw"
    assert not any(raw.iterdir())
