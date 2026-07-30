"""Générateurs de templates Excel starter pour un nouveau projet EthoFlow.

Appelés par `create_project.py` pour déposer un `<project>/<name>_sessions.xlsx`
à la racine du projet — l'utilisateur le remplit puis le passe à
`sync_from_excel.py`, qui détecte le schéma automatiquement depuis les
feuilles présentes.

Deux schémas selon le nombre d'animaux par vidéo :

- `single` : 1 souris par vidéo → 1 feuille `Sessions` à plat, avec
  `id` (nom du fichier vidéo, clé de session) en première colonne.
- `multi`  : N souris dans N arènes par vidéo → 3 feuilles
  (`Subjects`, `Trials_Videos`, `Arena_Mapping`).

Chaque template inclut une feuille `Instructions` qui explique quoi remplir
et donne la commande de sync. Colonnes taillées large, en-têtes en gras,
lignes d'exemple grisées pour montrer le format attendu.

Fonction publique unique :

    write_starter_excel(path, kind, project_name)
        → écrit path (xlsx), pas de retour, lève sur erreur d'écriture.
"""
from __future__ import annotations

from pathlib import Path


def _apply_common_style(ws, header_row: int = 1) -> None:
    """Met en gras la ligne d'en-têtes + freeze au-dessous."""
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[header_row]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # Freeze sous la ligne d'en-têtes
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _autosize(ws, min_w: int = 12, max_w: int = 40) -> None:
    """Ajuste la largeur des colonnes au contenu, avec clamp min/max."""
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            length = len(str(cell.value).split("\n")[0])
            if length > max_len:
                max_len = length
        w = max(min_w, min(max_len + 2, max_w))
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def _write_instructions_single(ws, project_name: str) -> None:
    from openpyxl.styles import Font, Alignment
    ws.append(["Projet :", project_name])
    ws.append(["Schéma :", "1 animal par vidéo (single)"])
    ws.append([])
    ws.append(["Feuille 'Sessions' — une ligne par VIDÉO (= une session)"])
    ws.append([])
    ws.append(["COLONNE OBLIGATOIRE"])
    ws.append(["  id"])
    ws.append(["      Nom du fichier vidéo SANS extension."])
    ws.append(["      Ex. id=970 → cherche 970.mp4 dans le dossier vidéos."])
    ws.append(["      C'est la clé unique de la session et le nom du dossier"])
    ws.append(["      créé dans data/raw/ (préfixé BV-)."])
    ws.append(["      Deux lignes ne peuvent pas avoir le même id."])
    ws.append([])
    ws.append(["COLONNES RECOMMANDÉES"])
    ws.append(["  mouse_id"])
    ws.append(["      Identifie l'ANIMAL. Peut se répéter sur plusieurs lignes"])
    ws.append(["      si la même souris est filmée à plusieurs timepoints"])
    ws.append(["      (design longitudinal) : id=970-M1 et id=970-M2 avec"])
    ws.append(["      mouse_id=970 dans les deux cas. Permet de regrouper les"])
    ws.append(["      sessions par animal dans les analyses."])
    ws.append(["  group"])
    ws.append(["      Variable de comparaison principale (ex. génotype,"])
    ws.append(["      traitement). C'est ce qui sépare tes groupes dans"])
    ws.append(["      analyze_vame.py."])
    ws.append([])
    ws.append(["TOUTES LES AUTRES COLONNES SONT LIBRES"])
    ws.append(["  Tu peux les renommer, les supprimer, ou en ajouter."])
    ws.append(["  Chaque colonne remplie est recopiée telle quelle dans le"])
    ws.append(["  metadata.yaml de la session, et devient utilisable comme"])
    ws.append(["  variable de groupement dans les analyses."])
    ws.append(["  Celles proposées ici (sex, cage, birth_date, genotype_*,"])
    ws.append(["  captopril...) sont des exemples adaptés au projet MCC —"])
    ws.append(["  adapte-les à ton étude."])
    ws.append(["  Laisse vide une cellule si l'info n'existe pas."])
    ws.append([])
    ws.append(["SYNC — depuis l'env conda 'ethoflow'"])
    ws.append(["  Mode interactif (le script demande ce qui manque) :"])
    ws.append(["      python scripts/sync_from_excel.py"])
    ws.append([])
    ws.append(["  Mode arguments :"])
    ws.append(["      python scripts/sync_from_excel.py \\"])
    ws.append(["          --project-dir <chemin de ce projet> \\"])
    ws.append(["          --videos-dir <dossier contenant les .mp4>"])
    ws.append([])
    ws.append(["  Répète pour chaque batch d'acquisition (--videos-dir change,"])
    ws.append(["  l'Excel reste le même). --overwrite pour re-générer une"])
    ws.append(["  metadata déjà existante, --dry-run pour prévisualiser."])
    # Style : en-têtes de section en gras (lignes en MAJUSCULES, "Feuille ...",
    # ou finissant par ":")
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=1)
        v = str(cell.value) if cell.value else ""
        if not v:
            continue
        is_section = (
            v.endswith(":")
            or v.startswith("Feuille")
            or (v == v.upper() and len(v) > 3 and not v.startswith(" "))
        )
        if is_section:
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 72


def _write_sessions_single(ws) -> None:
    headers = [
        "id",        # OBLIGATOIRE — nom du fichier vidéo, clé unique de session
        "mouse_id",  # RECOMMANDÉ — identifie l'animal (regroupement longitudinal)
        "group",     # RECOMMANDÉ — variable de comparaison principale
        # ── Colonnes libres : renomme / supprime / ajoute à volonté ──
        "date",  # date d'enregistrement, à remplir si tu veux la tracer
        "sex", "cage", "tail_label", "birth_date",
        "line", "origin",
        "genotype_mcc", "genotype_cdh5_cre", "genotype_col1_egfp",
        "captopril", "notes",
    ]
    ws.append(headers)
    # Trois lignes d'exemple grisées pour montrer le format attendu.
    # Les deux dernières montrent la MÊME souris (970) enregistrée à deux
    # timepoints : `id` diffère (donc 2 sessions), `mouse_id` est identique.
    example_rows = [
        ["971", 971, "MCCiECKO", "2026-06-08", "F", "CD330", 2, "2024-10-15",
         "MCC*Cdh5-cre", None, "fl/fl", "cre+", "+/+", "oui",
         "exemple à supprimer"],
        ["970-M1", 970, "MCCf/f", "2026-06-08", "F", "CD329", 1, "2024-10-15",
         "MCC*Cdh5-cre", None, "fl/fl", "cre+", "+/+", "oui",
         "exemple — même souris, timepoint 1"],
        ["970-M2", 970, "MCCf/f", "2026-09-14", "F", "CD329", 1, "2024-10-15",
         "MCC*Cdh5-cre", None, "fl/fl", "cre+", "+/+", "oui",
         "exemple — même souris, timepoint 2"],
    ]
    from openpyxl.styles import Font, PatternFill
    grey_font = Font(color="888888", italic=True)
    grey_fill = PatternFill("solid", fgColor="F2F2F2")
    for row in example_rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = grey_font
            cell.fill = grey_fill
    _apply_common_style(ws)
    _autosize(ws)


def _write_instructions_multi(ws, project_name: str) -> None:
    from openpyxl.styles import Font
    ws.append(["Projet :", project_name])
    ws.append(["Schéma :", "N animaux par vidéo (multi), typiquement 4 arènes"])
    ws.append([])
    ws.append(["COLONNES OBLIGATOIRES"])
    ws.append(["  Trials_Videos.TrialCode  — identifiant unique de la vidéo"])
    ws.append(["  Arena_Mapping.TrialCode  — référence vers Trials_Videos"])
    ws.append(["  Arena_Mapping.Arena      — numéro d'arène (1, 2, 3, 4)"])
    ws.append(["  Arena_Mapping.MouseID    — souris dans cette arène"])
    ws.append([])
    ws.append(["Feuille 'Subjects'"])
    ws.append(["  - Une ligne par souris (MouseID unique)."])
    ws.append(["  - Renseigne le groupe expérimental par timepoint."])
    ws.append(["  - Colonnes libres : adapte-les à ton étude."])
    ws.append([])
    ws.append(["Feuille 'Trials_Videos'"])
    ws.append(["  - Une ligne par vidéo enregistrée."])
    ws.append(["  - TrialCode conventionnel : OF-<M1|M2>-<YYYYMMDD>-V<##>"])
    ws.append(["  - Le fichier vidéo correspondant doit être dans le dossier"])
    ws.append(["    vidéos, nommé selon 'Original file name' ou <TrialCode>.mp4."])
    ws.append(["  - FPS / Width / Height sont optionnels (info seulement)."])
    ws.append([])
    ws.append(["Feuille 'Arena_Mapping'"])
    ws.append(["  - Une ligne par (vidéo × arène). N lignes = N vidéos × 4."])
    ws.append(["  - MouseID = souris présente dans cette arène pour cette vidéo."])
    ws.append(["  - Une même souris peut apparaître dans plusieurs vidéos"])
    ws.append(["    (timepoints différents) — c'est le cas longitudinal."])
    ws.append([])
    ws.append(["TOUTES LES AUTRES COLONNES SONT LIBRES"])
    ws.append(["  Renomme, supprime ou ajoute selon ton étude. Ce qui est"])
    ws.append(["  rempli finit dans le metadata.yaml et devient utilisable"])
    ws.append(["  comme variable de groupement dans les analyses."])
    ws.append([])
    ws.append(["SYNC — depuis l'env conda 'ethoflow'"])
    ws.append(["  Mode interactif (le script demande ce qui manque) :"])
    ws.append(["      python scripts/sync_from_excel.py"])
    ws.append([])
    ws.append(["  Mode arguments :"])
    ws.append(["      python scripts/sync_from_excel.py \\"])
    ws.append(["          --project-dir <chemin de ce projet> \\"])
    ws.append(["          --videos-dir <dossier contenant les .mp4>"])
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=1)
        v = str(cell.value) if cell.value else ""
        if not v:
            continue
        is_section = (
            v.endswith(":")
            or v.startswith("Feuille")
            or (v == v.upper() and len(v) > 3 and not v.startswith(" "))
        )
        if is_section:
            cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 72


def _write_subjects_multi(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    ws.append([
        "MouseID", "Baseline group (M1)", "ANGII group (M2)",
        "Stress (CUS?)", "Notes",
    ])
    example_rows = [
        [1, "CUS", "CUS+ANGII", "yes", "exemple à supprimer"],
        [11, "SHAM", "SHAM+ANGII", "no", "exemple à supprimer"],
    ]
    grey_font = Font(color="888888", italic=True)
    grey_fill = PatternFill("solid", fgColor="F2F2F2")
    for row in example_rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = grey_font
            cell.fill = grey_fill
    _apply_common_style(ws)
    _autosize(ws)


def _write_trials_multi(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    ws.append([
        "TrialCode", "Timepoint", "Date (YYYY-MM-DD)", "VideoNo",
        "Original file name", "FPS", "Width", "Height", "Notes",
    ])
    example = [
        "OF-M1-20260210-V01", "M1", "2026-02-10", "01",
        "V01.mp4", 25, 1280, 1024, "exemple à supprimer",
    ]
    ws.append(example)
    grey_font = Font(color="888888", italic=True)
    grey_fill = PatternFill("solid", fgColor="F2F2F2")
    for cell in ws[ws.max_row]:
        cell.font = grey_font
        cell.fill = grey_fill
    _apply_common_style(ws)
    _autosize(ws)


def _write_arena_mapping_multi(ws) -> None:
    from openpyxl.styles import Font, PatternFill
    ws.append([
        "ArenaCode", "TrialCode", "Timepoint", "Arena", "MouseID",
        "Notes",
    ])
    example_rows = [
        ["OF-M1-20260210-V01_A1", "OF-M1-20260210-V01", "M1", 1, 15,
         "exemple à supprimer"],
        ["OF-M1-20260210-V01_A2", "OF-M1-20260210-V01", "M1", 2, 16,
         "exemple à supprimer"],
        ["OF-M1-20260210-V01_A3", "OF-M1-20260210-V01", "M1", 3, 17,
         "exemple à supprimer"],
        ["OF-M1-20260210-V01_A4", "OF-M1-20260210-V01", "M1", 4, 18,
         "exemple à supprimer"],
    ]
    grey_font = Font(color="888888", italic=True)
    grey_fill = PatternFill("solid", fgColor="F2F2F2")
    for row in example_rows:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.font = grey_font
            cell.fill = grey_fill
    _apply_common_style(ws)
    _autosize(ws)


def write_starter_excel(path: Path, kind: str, project_name: str) -> None:
    """Écrit un template Excel starter pour un nouveau projet EthoFlow.

    Args:
        path: où écrire (ex : <project>/<project>_sessions.xlsx)
        kind: "single" (1 animal/vidéo) ou "multi" (N animaux/vidéo)
        project_name: nom du projet, affiché en tête de la feuille
            Instructions pour rappel visuel.

    Nécessite `openpyxl` (livré avec l'env `ethoflow`).
    """
    if kind not in ("single", "multi"):
        raise ValueError(f"kind doit être 'single' ou 'multi', reçu : {kind}")
    try:
        from openpyxl import Workbook
    except ImportError:
        raise RuntimeError(
            "openpyxl requis pour générer le template Excel. "
            "Active l'env ethoflow (`conda activate ethoflow`) puis relance."
        )

    wb = Workbook()
    # Renomme la feuille par défaut en Instructions
    inst = wb.active
    inst.title = "Instructions"

    if kind == "single":
        _write_instructions_single(inst, project_name)
        sessions = wb.create_sheet("Sessions")
        _write_sessions_single(sessions)
    else:  # multi
        _write_instructions_multi(inst, project_name)
        subjects = wb.create_sheet("Subjects")
        _write_subjects_multi(subjects)
        trials = wb.create_sheet("Trials_Videos")
        _write_trials_multi(trials)
        arenas = wb.create_sheet("Arena_Mapping")
        _write_arena_mapping_multi(arenas)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    # Petit test rapide en ligne de commande
    import argparse
    parser = argparse.ArgumentParser(
        description="Génère un template Excel starter pour un projet EthoFlow."
    )
    parser.add_argument("--path", required=True, type=Path)
    parser.add_argument("--kind", required=True, choices=["single", "multi"])
    parser.add_argument("--project-name", default="mon-projet")
    args = parser.parse_args()
    write_starter_excel(args.path, args.kind, args.project_name)
    print(f"✅ {args.path}")
